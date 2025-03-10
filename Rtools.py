import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import re
import time
import datetime
import sys
import os
import traceback

class ReportConverter:
    def __init__(self):
        # 检查时间锁
        if not self.check_time_lock():
            print("校验出错！！请检查程序版本！！")
            sys.exit(1)
            
        self.root = tk.Tk()
        self.root.title("报表转换工具 @ 2025")
        # 设置固定窗口大小
        self.root.geometry("1024x768")
        self.root.resizable(False, False)
        
        self.file_path = None
        self.workbook = None
        self.sheet_names = []
        
        # 添加期间数据存储变量
        self.period_data = {
            'balance_sheet': {},
            'cash_flow': {},
            'income_statement': {}
        }
        
        # 添加数据存储变量
        self.processed_data = {
            'balance_sheet': {},
            'cash_flow': {},
            'income_statement': {}
        }
        
        # 添加模板数据
        self.templates = {
            'balance_sheet': self.get_balance_sheet_template(),
            'cash_flow': self.get_cash_flow_template(),
            'income_statement': self.get_income_statement_template()
        }
        
        # 添加日志变量
        self.log_text = None
        
        self.setup_ui()
        
        # 显示欢迎信息
        self.log_message("=" * 50, "INFO")
        self.log_message("报表转换工具 @ 2025", "INFO")
        self.log_message("=" * 50, "INFO")
        self.log_message("准备就绪，请选择Excel文件开始处理...", "INFO")
    
    def setup_ui(self):
        """设置用户界面"""
        # 创建主框架
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # 创建主滚动框架
        self.main_canvas = tk.Canvas(main_frame, height=700)  # 设置固定高度
        self.scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=self.main_canvas.yview)
        self.scrollable_frame = ttk.Frame(self.main_canvas)
        
        # 配置滚动区域
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))
        )
        self.main_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw", width=980)  # 设置固定宽度
        self.main_canvas.configure(yscrollcommand=self.scrollbar.set)
        
        # 布局主滚动框架
        self.main_canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        
        # 创建文件选择按钮
        file_frame = ttk.Frame(self.scrollable_frame)
        file_frame.pack(fill="x", pady=5)
        
        self.select_file_btn = ttk.Button(
            file_frame, 
            text="选择Excel文件",
            command=self.select_file,
            width=20
        )
        self.select_file_btn.pack(side="left")
        
        # 创建sheet选择框架
        self.sheet_frame = ttk.LabelFrame(self.scrollable_frame, text="选择对应的Sheet")
        self.sheet_frame.pack(pady=5, fill="x")
        
        # 使用Grid布局管理sheet选择区域
        self.sheet_frame.grid_columnconfigure(1, weight=1)
        
        # 资产负债表sheet选择
        ttk.Label(self.sheet_frame, text="资产负债表:").grid(row=0, column=0, padx=5, pady=2, sticky="e")
        self.balance_sheet_var = tk.StringVar()
        self.balance_sheet_combo = ttk.Combobox(
            self.sheet_frame, 
            textvariable=self.balance_sheet_var,
            state="readonly",
            width=50
        )
        self.balance_sheet_combo.grid(row=0, column=1, padx=5, pady=2, sticky="w")
        
        # 现金流量表sheet选择
        ttk.Label(self.sheet_frame, text="现金流量表:").grid(row=1, column=0, padx=5, pady=2, sticky="e")
        self.cash_flow_var = tk.StringVar()
        self.cash_flow_combo = ttk.Combobox(
            self.sheet_frame, 
            textvariable=self.cash_flow_var,
            state="readonly",
            width=50
        )
        self.cash_flow_combo.grid(row=1, column=1, padx=5, pady=2, sticky="w")
        
        # 损益表sheet选择
        ttk.Label(self.sheet_frame, text="损益表:").grid(row=2, column=0, padx=5, pady=2, sticky="e")
        self.income_statement_var = tk.StringVar()
        self.income_statement_combo = ttk.Combobox(
            self.sheet_frame, 
            textvariable=self.income_statement_var,
            state="readonly",
            width=50
        )
        self.income_statement_combo.grid(row=2, column=1, padx=5, pady=2, sticky="w")
        
        # 添加确认按钮
        self.confirm_sheets_btn = ttk.Button(
            self.sheet_frame,
            text="确认选择并分析期间",
            command=self.analyze_periods,
            width=20
        )
        self.confirm_sheets_btn.grid(row=3, column=0, columnspan=2, pady=5)
        
        # 创建期间选择框架
        self.period_frame = ttk.LabelFrame(self.scrollable_frame, text="选择期间列")
        self.period_frame.pack(pady=5, fill="x")
        
        # 创建按钮和进度条框架
        bottom_frame = ttk.Frame(self.scrollable_frame)
        bottom_frame.pack(fill="x", pady=5)
        
        # 创建按钮框架
        self.button_frame = ttk.Frame(bottom_frame)
        self.button_frame.pack(side="left", padx=5)
        
        # 添加处理按钮
        self.process_btn = ttk.Button(
            self.button_frame,
            text="处理数据",
            command=self.process_data,
            state="disabled",
            width=15
        )
        self.process_btn.pack(side="left", padx=5)
        
        # 添加导出按钮
        self.export_btn = ttk.Button(
            self.button_frame,
            text="导出数据",
            command=self.export_data,
            state="disabled",
            width=15
        )
        self.export_btn.pack(side="left", padx=5)
        
        # 添加进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            bottom_frame,
            variable=self.progress_var,
            maximum=100,
            length=300
        )
        self.progress_bar.pack(side="right", padx=5)
        
        # 创建日志区域
        log_frame = ttk.LabelFrame(self.scrollable_frame, text="处理日志")
        log_frame.pack(fill="x", pady=5)
        
        # 创建日志文本框和滚动条
        log_scroll = ttk.Scrollbar(log_frame)
        log_scroll.pack(side="right", fill="y")
        
        self.log_text = tk.Text(
            log_frame,
            height=10,
            width=100,
            wrap=tk.WORD,
            yscrollcommand=log_scroll.set
        )
        self.log_text.pack(fill="both", expand=True)
        log_scroll.config(command=self.log_text.yview)
        
        # 配置日志文本框标签
        self.log_text.tag_config("INFO", foreground="black")
        self.log_text.tag_config("SUCCESS", foreground="green")
        self.log_text.tag_config("WARNING", foreground="orange")
        self.log_text.tag_config("ERROR", foreground="red")
        self.log_text.tag_config("DEBUG", foreground="blue")
        
        # 绑定鼠标滚轮事件
        self.main_canvas.bind("<MouseWheel>", self._on_mousewheel)
    
    def _on_mousewheel(self, event):
        """处理鼠标滚轮事件"""
        self.main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    
    def select_file(self):
        """选择Excel文件并读取sheet名称"""
        if not self.check_time_lock():
            self.log_message("校验出错！！请检查程序版本！！", "ERROR")
            return
            
        self.file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if self.file_path:
            try:
                # 检查文件扩展名
                file_extension = os.path.splitext(self.file_path)[1].lower()
                
                if file_extension == '.xls':
                    # 使用xlrd处理.xls文件
                    import xlrd
                    wb = xlrd.open_workbook(self.file_path, formatting_info=True)
                    # 将.xls转换为.xlsx
                    temp_path = self.file_path + 'x'  # 临时文件路径
                    self.convert_xls_to_xlsx(wb, temp_path)
                    self.file_path = temp_path
                
                # 使用openpyxl打开.xlsx文件
                # 先加载公式版本，用于获取公式值
                self.workbook_with_formulas = openpyxl.load_workbook(self.file_path)
                # 再加载数据版本，用于获取普通数据
                self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
                self.sheet_names = self.workbook.sheetnames
                
                for combo in [self.balance_sheet_combo, 
                            self.cash_flow_combo, 
                            self.income_statement_combo]:
                    combo['values'] = self.sheet_names
                
                # 添加文件名到日志
                file_name = os.path.basename(self.file_path)
                self.log_message(f"导入文件：{file_name}", "SUCCESS")
                
            except Exception as e:
                self.log_message(f"文件加载失败：{str(e)}", "ERROR")
    
    def convert_xls_to_xlsx(self, wb, save_path):
        """将.xls文件转换为.xlsx格式"""
        new_wb = openpyxl.Workbook()
        
        for sheet_name in wb.sheet_names():
            # 获取原始sheet
            sheet = wb.sheet_by_name(sheet_name)
            # 创建新sheet
            new_sheet = new_wb.create_sheet(title=sheet_name)
            
            # 复制数据
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    cell_value = sheet.cell_value(row, col)
                    new_sheet.cell(row=row+1, column=col+1, value=cell_value)
        
        # 删除默认创建的sheet
        if 'Sheet' in new_wb.sheetnames:
            new_wb.remove(new_wb['Sheet'])
            
        # 保存为.xlsx格式
        new_wb.save(save_path)
    
    def find_period_columns(self, sheet):
        """查找包含期间信息的列"""
        keywords = ['本期', '上期', '同期', '年初', '期初', '期末', '年', '月', '季', '期', '/', '-']
        period_columns = {}
        
        # 限制检查的最大列数
        max_cols = min(sheet.max_column, 100)  # 最多检查100列
        
        # 检查前7行
        for row in range(1, 8):
            for col in range(1, max_cols + 1):
                # 每处理20列更新一次UI
                if col % 20 == 0:
                    self.root.update_idletasks()
                    
                cell_value = str(sheet.cell(row, col).value or '')
                # 优化：先检查单元格是否为空
                if not cell_value:
                    continue
                    
                # 优化：使用任何关键词匹配
                if any(keyword in cell_value for keyword in keywords):
                    col_letter = get_column_letter(col)
                    period_columns[col_letter] = cell_value
                    self.log_message(f"找到期间列: {col_letter} - {cell_value}", "DEBUG")
        
        # 如果找不到期间列，记录警告
        if not period_columns:
            self.log_message(f"警告：在工作表 {sheet.title} 中未找到任何期间列", "WARNING")
            
        return period_columns
    
    def analyze_periods(self):
        """分析所选sheet中的期间列"""
        if not self.check_time_lock():
            self.log_message("校验出错！！请检查程序版本！！", "ERROR")
            return
            
        try:
            if not self.workbook:
                self.log_message("请先选择Excel文件", "ERROR")
                return
                
            if not all([self.balance_sheet_var.get(),
                       self.cash_flow_var.get(),
                       self.income_statement_var.get()]):
                self.log_message("请先选择所有需要的sheet！", "WARNING")
                return
            
            # 禁用按钮，防止重复点击
            self.confirm_sheets_btn['state'] = 'disabled'
            self.log_message("开始分析期间，请稍候...", "INFO")
            
            # 创建进度条
            progress_window = tk.Toplevel(self.root)
            progress_window.title("处理中")
            progress_window.geometry("300x100")
            progress_window.transient(self.root)
            progress_window.grab_set()
            
            progress_label = ttk.Label(progress_window, text="正在分析期间列，请稍候...")
            progress_label.pack(pady=10)
            
            progress_bar = ttk.Progressbar(progress_window, mode="indeterminate")
            progress_bar.pack(fill="x", padx=20, pady=10)
            progress_bar.start()
            
            # 使用after方法延迟执行，让UI能够更新
            def process_task():
                try:
                    # 清空之前的期间选择框
                    for widget in self.period_frame.winfo_children():
                        widget.destroy()
                        
                    # 分析并显示各个sheet的期间列
                    sheets = {
                        'balance_sheet': self.balance_sheet_var.get(),
                        'cash_flow': self.cash_flow_var.get(),
                        'income_statement': self.income_statement_var.get()
                    }
                    
                    # 先找到所有期间列
                    periods_data = {}
                    for sheet_type, sheet_name in sheets.items():
                        progress_label.config(text=f"正在分析 {sheet_name} 的期间列...")
                        progress_window.update()
                        
                        sheet = self.workbook[sheet_name]
                        periods = self.find_period_columns(sheet)
                        periods_data[sheet_type] = periods
                        
                        # 更新UI
                        self.root.update_idletasks()
                    
                    # 然后进行预处理
                    for sheet_type, sheet_name in sheets.items():
                        progress_label.config(text=f"正在预处理 {sheet_name}...")
                        progress_window.update()
                        
                        sheet = self.workbook[sheet_name]
                        self.preprocess_sheet(sheet, sheet_name)
                        
                        # 更新UI
                        self.root.update_idletasks()
                    
                    # 使用之前保存的期间信息创建UI
                    progress_label.config(text="正在创建期间选择界面...")
                    progress_window.update()
                    
                    row = 0
                    for sheet_type, sheet_name in sheets.items():
                        periods = periods_data[sheet_type]
                        
                        # 创建期间选择区域
                        ttk.Label(self.period_frame, text=f"{sheet_name}:").grid(
                            row=row, column=0, columnspan=2, padx=5, pady=5, sticky="w")
                        
                        # 本期选择
                        ttk.Label(self.period_frame, text="本期:").grid(
                            row=row+1, column=0, padx=5, pady=2, sticky="e")
                        current_period = ttk.Combobox(
                            self.period_frame,
                            values=list(periods.values()),
                            state="readonly",
                            width=40
                        )
                        current_period.grid(row=row+1, column=1, padx=5, pady=2, sticky="w")
                        
                        # 上期选择
                        ttk.Label(self.period_frame, text="上期:").grid(
                            row=row+2, column=0, padx=5, pady=2, sticky="e")
                        prev_period = ttk.Combobox(
                            self.period_frame,
                            values=list(periods.values()),
                            state="readonly",
                            width=40
                        )
                        prev_period.grid(row=row+2, column=1, padx=5, pady=2, sticky="w")
                        
                        # 年初选择
                        ttk.Label(self.period_frame, text="年初:").grid(
                            row=row+3, column=0, padx=5, pady=2, sticky="e")
                        year_start = ttk.Combobox(
                            self.period_frame,
                            values=list(periods.values()),
                            state="readonly",
                            width=40
                        )
                        year_start.grid(row=row+3, column=1, padx=5, pady=2, sticky="w")
                        
                        # 存储期间选择控件
                        self.period_data[sheet_type] = {
                            'current': current_period,
                            'previous': prev_period,
                            'year_start': year_start,
                            'columns': periods
                        }
                        
                        row += 4
                    
                    # 启用处理按钮
                    self.process_btn['state'] = 'normal'
                    self.confirm_sheets_btn['state'] = 'normal'
                    self.log_message("期间分析完成，可以开始处理数据", "SUCCESS")
                    
                    # 关闭进度窗口
                    progress_window.destroy()
                    
                except Exception as e:
                    self.log_message(f"分析期间时出错：{str(e)}", "ERROR")
                    progress_window.destroy()
                    self.confirm_sheets_btn['state'] = 'normal'
            
            # 使用after方法延迟执行，让UI能够更新
            self.root.after(100, process_task)
            
        except Exception as e:
            self.log_message(f"分析期间时出错：{str(e)}", "ERROR")
            self.confirm_sheets_btn['state'] = 'normal'
    
    def preprocess_sheet(self, sheet, sheet_name):
        """预处理工作表，删除包含'TB.global.'的列"""
        self.log_message(f"开始预处理工作表 {sheet_name}...", "INFO")
        
        # 获取工作表的最大列数
        max_col = sheet.max_column
        cols_to_delete = []
        
        # 设置最大检查列数，避免处理过大的表格
        max_cols_to_check = min(max_col, 100)  # 最多检查100列
        
        # 检查前20行，从后向前遍历列（最多检查100列）
        for col in range(max_col, max(1, max_col - max_cols_to_check), -1):
            # 每处理10列更新一次UI
            if col % 10 == 0:
                self.root.update_idletasks()
                
            col_letter = get_column_letter(col)
            found_tb_global = False
            
            # 检查前20行
            for row in range(1, min(21, sheet.max_row + 1)):
                cell_value = str(sheet[f"{col_letter}{row}"].value or "").strip().lower()
                # 扩大搜索范围，包含更多可能的变体
                if any(keyword in cell_value for keyword in ['tb', 'trial balance', 'global']):
                    found_tb_global = True
                    self.log_message(
                        f"在工作表 {sheet_name} 的第 {row} 行 {col_letter} 列找到匹配项", 
                        "DEBUG"
                    )
                    break
            
            if found_tb_global:
                cols_to_delete.append(col)
        
        # 如果要删除的列太多，给用户一个警告
        if len(cols_to_delete) > 20:
            self.log_message(f"警告：将删除 {len(cols_to_delete)} 列，这可能需要一些时间", "WARNING")
        
        # 批量删除列，而不是一次删除一列
        # 从后向前删除列，按照连续的块进行删除以提高效率
        if cols_to_delete:
            # 按照从大到小排序
            cols_to_delete.sort(reverse=True)
            
            # 分组连续的列
            groups = []
            current_group = [cols_to_delete[0]]
            
            for i in range(1, len(cols_to_delete)):
                if cols_to_delete[i] == cols_to_delete[i-1] - 1:
                    # 连续的列
                    current_group.append(cols_to_delete[i])
                else:
                    # 不连续，开始新的组
                    groups.append(current_group)
                    current_group = [cols_to_delete[i]]
            
            # 添加最后一组
            if current_group:
                groups.append(current_group)
            
            # 删除每组连续的列
            for group in groups:
                # 获取连续列的起始和结束
                start_col = min(group)
                count = len(group)
                
                # 删除连续的列
                sheet.delete_cols(start_col, count)
                self.log_message(f"在工作表 {sheet_name} 中删除了从第 {get_column_letter(start_col)} 列开始的 {count} 列", "INFO")
                
                # 更新UI
                self.root.update_idletasks()
        else:
            self.log_message(f"工作表 {sheet_name} 中没有找到需要删除的列", "INFO")
    
    def process_data(self):
        """处理各个报表数据"""
        if not self.check_time_lock():
            self.log_message("校验出错！！请检查程序版本！！", "ERROR")
            return
            
        try:
            self.progress_var.set(0)
            self.log_message("开始处理数据...", "INFO")
            
            # 处理各个报表
            self.process_balance_sheet()
            self.process_cash_flow()
            self.process_income_statement()
            
            self.progress_var.set(100)
            self.log_message("数据处理完成！", "SUCCESS")
            
            # 启用导出按钮
            self.export_btn['state'] = 'normal'
            
        except Exception as e:
            self.log_message(f"处理数据时出错：{str(e)}", "ERROR")
    
    def process_balance_sheet(self):
        """处理资产负债表数据"""
        sheet_name = self.balance_sheet_var.get()
        sheet = self.workbook[sheet_name]
        data = self.period_data['balance_sheet']
        template = self.templates['balance_sheet'].copy()
        
        # 获取选中的列
        columns = self.get_period_columns(data)
        
        # 遍历所有行
        matched_items = set()  # 记录已匹配的项目
        for row in range(1, sheet.max_row + 1):
            self.progress_var.set((row / sheet.max_row) * 100)
            self.root.update_idletasks()
            
            # 处理左侧（资产部分）
            left_item = str(sheet.cell(row, 1).value or '').strip()
            if left_item:
                self.process_balance_sheet_item(left_item, row, columns, template, matched_items)
            
            # 处理右侧（负债和所有者权益部分）
            # 通常在第5列或第6列开始
            for col in range(5, 7):  # 尝试这两列
                right_item = str(sheet.cell(row, col).value or '').strip()
                if right_item:
                    # 获取右侧数据的列偏移
                    col_offset = col - 1
                    right_columns = self.adjust_columns(columns, col_offset)
                    self.process_balance_sheet_item(right_item, row, right_columns, template, matched_items)
                    break
        
        # 打印未匹配的项目（用于调试）
        unmatched_items = set(template.keys()) - matched_items
        if unmatched_items:
            print("未匹配的项目：", unmatched_items)
        
        self.processed_data['balance_sheet'] = template
        self.calculate_totals(template, 'balance_sheet')
    
    def process_balance_sheet_item(self, item_name, row, columns, template, matched_items):
        """处理资产负债表单个项目"""
        if not item_name:
            return
            
        # 在模板中查找匹配项
        for template_name in template.keys():
            if template_name not in matched_items and self.match_item_name(item_name, template_name):
                # 获取各期数据
                values = self.get_period_values(self.workbook[self.balance_sheet_var.get()], 
                                             row, columns)
                template[template_name].update(values)
                matched_items.add(template_name)
                break
    
    def adjust_columns(self, columns, offset):
        """调整列索引以适应右侧数据"""
        adjusted_columns = {}
        for period, col in columns.items():
            if col:
                # 将列字母转换为数字，加上偏移量，再转回列字母
                col_num = self.get_column_index(col)
                new_col = get_column_letter(col_num + offset)
                adjusted_columns[period] = new_col
            else:
                adjusted_columns[period] = None
        return adjusted_columns
    
    def process_cash_flow(self):
        """处理现金流量表数据"""
        sheet_name = self.cash_flow_var.get()
        sheet = self.workbook[sheet_name]
        data = self.period_data['cash_flow']
        template = self.templates['cash_flow'].copy()
        
        # 获取选中的列
        columns = self.get_period_columns(data)
        
        # 遍历所有行
        for row in range(1, sheet.max_row + 1):
            self.progress_var.set((row / sheet.max_row) * 100)
            self.root.update_idletasks()
            
            item_name = str(sheet.cell(row, 1).value or '').strip()
            if not item_name:
                continue
                
            # 在模板中查找匹配项
            for template_name in template.keys():
                if self.match_item_name(item_name, template_name):
                    # 获取各期数据
                    values = self.get_period_values(sheet, row, columns)
                    template[template_name].update(values)
                    break
        
        self.processed_data['cash_flow'] = template
        
        self.calculate_totals(template, 'cash_flow')
    
    def process_income_statement(self):
        """处理损益表数据"""
        sheet_name = self.income_statement_var.get()
        sheet = self.workbook[sheet_name]
        data = self.period_data['income_statement']
        template = self.templates['income_statement'].copy()
        
        # 获取选中的列
        columns = self.get_period_columns(data)
        
        # 遍历所有行
        for row in range(1, sheet.max_row + 1):
            self.progress_var.set((row / sheet.max_row) * 100)
            self.root.update_idletasks()
            
            item_name = str(sheet.cell(row, 1).value or '').strip()
            if not item_name:
                continue
                
            # 在模板中查找匹配项
            for template_name in template.keys():
                if self.match_item_name(item_name, template_name):
                    # 获取各期数据
                    values = self.get_period_values(sheet, row, columns)
                    template[template_name].update(values)
                    break
        
        self.processed_data['income_statement'] = template
        
        self.calculate_totals(template, 'income_statement')
    
    def get_column_letter(self, header_value, columns):
        """根据列标题获取列字母"""
        for col_letter, value in columns.items():
            if value == header_value:
                return col_letter
        return None
    
    def get_column_index(self, column_letter):
        """将列字母转换为列索引"""
        result = 0
        for i, c in enumerate(reversed(column_letter.upper())):
            result += (ord(c) - ord('A') + 1) * (26 ** i)
        return result
    
    def export_data(self):
        """导出数据到Excel"""
        if not self.check_time_lock():
            self.log_message("校验出错！！请检查程序版本！！", "ERROR")
            return
            
        try:
            workbook = openpyxl.Workbook()
            self.log_message("开始导出数据...", "INFO")
            
            # 导出各个报表数据
            self.export_sheet(workbook, "资产负债表", self.processed_data['balance_sheet'])
            self.export_sheet(workbook, "现金流量表", self.processed_data['cash_flow'])
            self.export_sheet(workbook, "损益表", self.processed_data['income_statement'])
            
            # 计算并导出财务指标
            self.log_message("计算财务指标...", "INFO")
            indicators = self.calculate_financial_indicators(self.processed_data['balance_sheet'], self.processed_data['income_statement'], self.processed_data['cash_flow'])
            self.export_financial_indicators(workbook, indicators)
            
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            if save_path:
                workbook.save(save_path)
                # 添加导出文件名到日志
                file_name = os.path.basename(save_path)
                self.log_message(f"数据已导出到：{file_name}", "SUCCESS")
        except Exception as e:
            self.log_message(f"导出失败：{str(e)}", "ERROR")
    
    def export_sheet(self, workbook, sheet_name, data):
        """导出单个sheet的数据"""
        ws = workbook.create_sheet(sheet_name)
        
        # 设置表头样式
        header_style = openpyxl.styles.NamedStyle(name='header')
        header_style.font = openpyxl.styles.Font(bold=True)
        header_style.fill = openpyxl.styles.PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # 写入表头
        headers = ["科目名称", "行次", "本期", "上期", "年初"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(1, col, header)
            cell.style = header_style
        
        # 写入数据
        for row, (item_name, values) in enumerate(data.items(), start=2):
            ws.cell(row, 1, item_name)
            ws.cell(row, 2, values['行次'])
            ws.cell(row, 3, values['本期'])
            ws.cell(row, 4, values['上期'])
            ws.cell(row, 5, values['年初'])
        
        # 调整列宽
        self.adjust_column_width(ws)
    
    def get_balance_sheet_template(self):
        """获取资产负债表模板"""
        template = [
            ('流动资产：', 1),
            ('    货币资金', 2),
            ('    交易性金融资产', 3),
            ('    应收票据', 4),
            ('    应收账款', 5),
            ('    预付款项', 6),
            ('    应收利息', 7),
            ('    应收股利', 8),
            ('    其他应收款', 9),
            ('    存货', 10),
            ('    一年内到期的非流动资产', 11),
            ('    其他流动资产', 12),
            ('    待摊费用', 13),
            ('流动资产合计', 14),
            ('非流动资产：', 15),
            ('    可供出售金融资产', 16),
            ('    持有至到期投资', 17),
            ('    长期应收款', 18),
            ('    长期股权投资', 19),
            ('    投资性房地产', 20),
            ('    固定资产', 21),
            ('    在建工程', 22),
            ('    工程物资', 23),
            ('    固定资产清理', 24),
            ('    生产性生物资产', 25),
            ('    油气资产', 26),
            ('    无形资产', 27),
            ('    开发支出', 28),
            ('    商誉', 29),
            ('    长期待摊费用', 30),
            ('    递延所得税资产', 31),
            ('    其他非流动资产', 32),
            ('    其它长期资产', 33),
            ('非流动资产合计', 34),
            ('资产总计', 35),
            ('流动负债：', 36),
            ('    短期借款', 37),
            ('    交易性金融负债', 38),
            ('    应付票据', 39),
            ('    应付账款', 40),
            ('    预收款项', 41),
            ('    应付职工薪酬', 42),
            ('    应交税费', 43),
            ('    应付利息', 44),
            ('    应付股利', 45),
            ('    其他应付款', 46),
            ('    预提费用', 47),
            ('    一年内到期的非流动负债', 48),
            ('    其他流动负债', 49),
            ('流动负债合计', 50),
            ('非流动负债：', 51),
            ('    长期借款', 52),
            ('    应付债券', 53),
            ('    长期应付款', 54),
            ('    专项应付款', 55),
            ('    预计负债', 56),
            ('    递延所得税负债', 57),
            ('    其他非流动负债', 58),
            ('非流动负债合计', 59),
            ('负债合计', 60),
            ('所有者权益（或股东权益）：', 61),
            ('    股本', 62),
            ('    资本公积', 63),
            ('    减：库存股', 64),
            ('    盈余公积', 65),
            ('    未分配利润', 66),
            ('    少数股东权益', 67),
            ('    未结清对外担保余额', 68),
            ('所有者权益（或股东权益）合计', 69),
            ('负债和所有者权益（或股东权益）总计', 70)
        ]
        return {item[0]: {'行次': item[1], '本期': 0, '上期': 0, '年初': 0} for item in template}
    
    def get_cash_flow_template(self):
        """获取现金流量表模板"""
        template = [
            ('一、经营活动产生的现金流量:', 1),
            ('    销售商品、提供劳务收到的现金', 2),
            ('    收到的税费返还', 3),
            ('    收到其他与经营活动有关的现金', 4),
            ('    经营活动现金流入小计', 5),
            ('    购买商品、接受劳务支付的现金', 6),
            ('    支付给职工以及为职工支付的现金', 7),
            ('    支付的各项税费', 8),
            ('    支付其他与经营活动有关的现金', 9),
            ('    经营活动现金流出小计', 10),
            ('    经营活动产生的现金流量净额', 11),
            ('二、投资活动产生的现金流量:', 12),
            ('    收回投资收到的现金', 13),
            ('    取得投资收益收到的现金', 14),
            ('    处置固定资产、无形资产和其他长期资产收回的现金净额', 15),
            ('    处置子公司及其他营业单位收到的现金净额', 16),
            ('    收到其他与投资活动有关的现金', 17),
            ('    投资活动现金流入小计', 18),
            ('    购建固定资产、无形资产和其他长期资产支付的现金', 19),
            ('    投资支付的现金', 20),
            ('    取得子公司及其他营业单位支付的现金净额', 21),
            ('    支付其他与投资活动有关的现金', 22),
            ('    投资活动现金流出小计', 23),
            ('    投资活动产生的现金流量净额', 24),
            ('三、筹资活动产生的现金流量:', 25),
            ('    吸收投资收到的现金', 26),
            ('    取得借款收到的现金', 27),
            ('    收到其他与筹资活动有关的现金', 28),
            ('    筹资活动现金流入小计', 29),
            ('    偿还债务支付的现金', 30),
            ('    分配股利、利润或偿付利息支付的现金', 31),
            ('    支付其他与筹资活动有关的现金', 32),
            ('    筹资活动现金流出小计', 33),
            ('    筹资活动产生的现金流量净额', 34),
            ('四、汇率变动对现金及现金等价物的影响', 35),
            ('五、现金及现金等价物增加额', 36)
        ]
        return {item[0]: {'行次': item[1], '本期': 0, '上期': 0, '年初': 0} for item in template}
    
    def get_income_statement_template(self):
        """获取损益表模板"""
        template = [
            ('一、营业总收入', 1),
            ('    减：营业成本', 2),
            ('    营业税金及附加', 3),
            ('    销售费用', 4),
            ('    管理费用', 5),
            ('    财务费用（收益以"－"号填列）', 6),
            ('    资产减值损失', 7),
            ('    加：公允价值变动净收益（净损失以"－"号填列）', 8),
            ('    投资收益（净损失以"－"号填列）', 9),
            ('    其中：对联营企业和合营企业的投资收益', 10),
            ('二、营业利润（亏损以"－"填列）', 11),
            ('    加：营业外收入', 12),
            ('    减：营业外支出', 13),
            ('    其中：非流动资产处置净损失（净收益以"-"号填列）', 14),
            ('三、利润总额（亏损总额以"－"填列）', 15),
            ('    减：所得税费用', 16),
            ('四、净利润（净亏损以"－"号填列）', 17),
            ('五、每股收益', 18),
            ('    （一）基本每股收益', 19),
            ('    （二）稀释每股收益', 20)
        ]
        return {item[0]: {'行次': item[1], '本期': 0, '上期': 0, '年初': 0} for item in template}
    
    def match_item_name(self, source_name, template_name):
        """匹配项目名称"""
        if not source_name or not template_name:
            return False
            
        # 清理和标准化名称
        source_name = self.clean_item_name(source_name)
        template_name = self.clean_item_name(template_name)
        
        # 如果清理后为空，返回False
        if not source_name or not template_name:
            return False
            
        # 标准化处理
        source_name = source_name.lower()
        template_name = template_name.lower()
        
        # 移除前缀空格和冒号
        source_name = source_name.lstrip().lstrip('    ').rstrip(':：')
        template_name = template_name.lstrip().lstrip('    ').rstrip(':：')
        
        # 直接匹配
        if source_name == template_name:
            return True
            
        # 同义词匹配
        synonyms = {
            # 资产类
            '流动资产': ['流动资产', '流动资产：', '流动资产合计', '流动资产总计'],
            '非流动资产': ['非流动资产', '非流动资产：', '非流动资产合计', '非流动资产总计'],
            '资产总计': ['资产总计', '资产合计', '资产总额'],
            
            # 负债类
            '流动负债': ['流动负债', '流动负债：', '流动负债合计', '流动负债总计'],
            '非流动负债': ['非流动负债', '非流动负债：', '非流动负债合计', '非流动负债总计'],
            '负债合计': ['负债合计', '负债总计', '负债总额'],
            
            # 所有者权益类
            '所有者权益': ['所有者权益', '所有者权益（或股东权益）', '所有者权益（或股东权益）：', '股东权益'],
            '所有者权益合计': ['所有者权益合计', '所有者权益（或股东权益）合计', '股东权益合计', '所有者权益总计'],
            '负债和所有者权益总计': ['负债和所有者权益总计', '负债和所有者权益（或股东权益）总计', '负债及所有者权益总计', '负债和股东权益总计'],
            
            # 具体项目
            '预提费用': ['预提费用', '预提成本费用', '预提支出'],
            '应付股利': ['应付股利', '应付股息', '应付利息及应付股利'],
            '递延所得税负债': ['递延所得税负债', '递延税负债', '递延所得税'],
            '应付票据': ['应付票据', '应付汇票', '应付票据及应付账款'],
            '资本公积': ['资本公积', '资本公积金', '资本溢价'],
            '其他应付款': ['其他应付款', '其它应付款', '其他应付'],
            '一年内到期的非流动负债': ['一年内到期的非流动负债', '一年内到期非流动负债', '一年内到期长期负债'],
            '其他长期资产': ['其他长期资产', '其它长期资产', '其他非流动资产'],
            '应付债券': ['应付债券', '债券', '应付债券净额'],
            '固定资产清理': ['固定资产清理', '固定资产清算', '资产清理'],
            '专项应付款': ['专项应付款', '专项款', '专项应付'],
            '应付账款': ['应付账款', '应付款项', '应付票据及应付账款'],
            '长期借款': ['长期借款', '长期贷款', '长期债务'],
            '应付职工薪酬': ['应付职工薪酬', '应付工资', '工资福利', '应付工资薪酬'],
            '少数股东权益': ['少数股东权益', '少数股东', '少数股东权益合计'],
            '其他流动负债': ['其他流动负债', '其它流动负债', '其他流动'],
            '交易性金融负债': ['交易性金融负债', '以公允价值计量且其变动计入当期损益的金融负债', '交易性负债'],
            '持有至到期投资': ['持有至到期投资', '持有到期投资', '持有至到期'],
            '应交税费': ['应交税费', '应交税金', '应缴税金', '应交税款'],
            '未结清对外担保余额': ['未结清对外担保余额', '对外担保余额', '担保余额'],
            '其他非流动负债': ['其他非流动负债', '其它非流动负债', '其他长期负债'],
            '短期借款': ['短期借款', '短期贷款', '短期债务'],
            '股本': ['股本', '实收资本', '实收资本(或股本)', '注册资本'],
            '应付利息': ['应付利息', '应付利息费用', '应付利息及应付股利'],
            '可供出售金融资产': ['可供出售金融资产', '可供出售的金融资产', '可供出售投资'],
            '盈余公积': ['盈余公积', '盈余公积金', '法定盈余'],
            '未分配利润': ['未分配利润', '未分配利润(未弥补亏损)', '留存收益', '累计利润'],
            '长期待摊费用': ['长期待摊费用', '待摊费用', '长期待摊', '待摊'],
            '预计负债': ['预计负债', '预计债务', '预提负债'],
            '长期应付款': ['长期应付款', '长期应付款项', '长期应付'],
            '预收款项': ['预收款项', '预收账款', '预收款', '合同负债'],
            '库存股': ['库存股', '减：库存股', '库存股份'],
            '工程物资': ['工程物资', '工程材料', '工程用料']
        }
        
        # 检查同义词
        for standard, variants in synonyms.items():
            if source_name in variants or template_name in variants:
                if source_name in variants and template_name in variants:
                    return True
                if source_name == standard and template_name in variants:
                    return True
                if template_name == standard and source_name in variants:
                    return True
        
        return False

    def clean_item_name(self, name):
        """清理项目名称"""
        if not name:
            return ""
            
        # 转换为字符串
        name = str(name)
        
        # 移除所有空格（包括前导空格、尾随空格和中间空格）
        name = ''.join(name.split())
        
        # 移除特殊字符
        chars_to_remove = [
            ':', '：', '(', ')', '（', '）', '、', '，', ',', 
            '；', ';', '"', '"', '"', ''', ''', '［', '］',
            '[', ']', '【', '】', '｛', '｝', '{', '}',
            '…', '......', '..', '。', '='
        ]
        for char in chars_to_remove:
            name = name.replace(char, '')
            
        # 统一中文字符
        name = name.replace('－', '-')
        name = name.replace('—', '-')
        name = name.replace('－', '-')
        name = name.replace('＋', '+')
        name = name.replace('／', '/')
        
        # 移除前导的层级标记和空格
        name = re.sub(r'^[\s\d\.]+', '', name)  # 移除前导数字和点
        name = re.sub(r'^[一二三四五六七八九十]+[、\s.]', '', name)  # 移除前导中文数字
        name = name.lstrip('    ').strip()
        
        return name.strip()

    def standardize_name(self, name):
        """标准化项目名称"""
        # 移除前导空格和层级标记
        name = name.lstrip().lstrip('    ').strip()
        
        # 移除常见的前缀和后缀
        prefixes = ['合计', '小计', '合计：', '小计：', '：', ':', '总计', '总额']
        for prefix in prefixes:
            if name.startswith(prefix):
                name = name[len(prefix):]
            if name.endswith(prefix):
                name = name[:-len(prefix)]
                
        # 统一括号内的内容
        name = re.sub(r'［.*?］', '', name)
        name = re.sub(r'\[.*?\]', '', name)
        name = re.sub(r'【.*?】', '', name)
        name = re.sub(r'\(.*?\)', '', name)
        name = re.sub(r'（.*?）', '', name)
        
        # 移除其他特殊标记
        name = name.replace('…', '')
        name = name.replace('——', '')
        name = name.replace('--', '')
        
        # 处理特殊的前缀
        name = re.sub(r'^[一二三四五六七八九十]+、', '', name)
        name = re.sub(r'^[0-9]+、', '', name)
        name = re.sub(r'^[A-Za-z]+、', '', name)
        
        # 移除"减："前缀
        if name.startswith('减：'):
            name = name[2:]
            
        return name.strip()
    
    def get_period_columns(self, data):
        """获取期间列信息"""
        return {
            '本期': self.get_column_letter(data['current'].get(), data['columns']) if data['current'].get() else None,
            '上期': self.get_column_letter(data['previous'].get(), data['columns']) if data['previous'].get() else None,
            '年初': self.get_column_letter(data['year_start'].get(), data['columns']) if data['year_start'].get() else None
        }
    
    def get_period_values(self, sheet, row, columns):
        """获取各期数据"""
        values = {'本期': 0, '上期': 0, '年初': 0}
        
        for period, col in columns.items():
            if col:
                cell = sheet.cell(row, self.get_column_index(col))
                value = cell.value
                
                # 如果值为None或0，尝试从带公式的工作簿获取值
                if (value is None or value == 0) and hasattr(self, 'workbook_with_formulas'):
                    try:
                        # 获取相同位置的单元格，但从带公式的工作簿中
                        sheet_name = sheet.title
                        formula_sheet = self.workbook_with_formulas[sheet_name]
                        formula_cell = formula_sheet.cell(row, self.get_column_index(col))
                        
                        # 如果单元格有公式，尝试计算公式结果
                        if formula_cell.data_type == 'f':
                            self.log_message(f"检测到公式单元格: {formula_cell.coordinate}, 公式: {formula_cell.value}", "INFO")
                            
                            # 尝试从公式中提取数值
                            if formula_cell.value and formula_cell.value.startswith('='):
                                # 记录原始公式
                                self.log_message(f"原始公式: {formula_cell.value}", "INFO")
                                
                                # 尝试获取计算结果
                                try:
                                    # 尝试使用内部值
                                    if hasattr(formula_cell, 'internal_value') and formula_cell.internal_value:
                                        value = formula_cell.internal_value
                                    # 或者尝试使用缓存值
                                    elif hasattr(formula_cell, '_value') and formula_cell._value:
                                        value = formula_cell._value
                                except Exception as e:
                                    self.log_message(f"无法计算公式结果: {str(e)}", "WARNING")
                    except Exception as e:
                        self.log_message(f"尝试从带公式的工作簿获取值时出错: {str(e)}", "WARNING")
                
                # 尝试转换为数值
                try:
                    # 处理字符串形式的数字（可能包含逗号等格式）
                    if isinstance(value, str):
                        value = value.replace(',', '')
                    values[period] = float(value) if value is not None else 0
                except (ValueError, TypeError):
                    values[period] = 0
                
                # 记录获取的值
                col_idx = self.get_column_index(col)
                col_letter = get_column_letter(col_idx)
                self.log_message(f"单元格 {sheet.title}!{col_letter}{row} 获取到的值: {values[period]}", "INFO")
        
        return values
    
    def adjust_column_width(self, ws):
        """调整列宽"""
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def calculate_totals(self, template, sheet_type):
        """计算合计项"""
        if sheet_type == 'balance_sheet':
            # 计算流动资产合计
            current_assets = sum(template[name]['本期'] for name in [
                '    货币资金', '    交易性金融资产', '    应收票据', '    应收账款',
                '    预付款项', '    应收利息', '    应收股利', '    其他应收款',
                '    存货', '    一年内到期的非流动资产', '    其他流动资产', '    待摊费用'
            ])
            template['流动资产合计'].update({'本期': current_assets})
            
            # 计算非流动资产合计
            non_current_assets = sum(template[name]['本期'] for name in [
                '    可供出售金融资产', '    持有至到期投资', '    长期应收款',
                '    长期股权投资', '    投资性房地产', '    固定资产', '    在建工程',
                '    工程物资', '    固定资产清理', '    生产性生物资产', '    油气资产',
                '    无形资产', '    开发支出', '    商誉', '    长期待摊费用',
                '    递延所得税资产', '    其他非流动资产', '    其它长期资产'
            ])
            template['非流动资产合计'].update({'本期': non_current_assets})
            
            # 计算资产总计
            template['资产总计'].update({'本期': current_assets + non_current_assets})
            
            # 计算流动负债合计
            current_liabilities = sum(template[name]['本期'] for name in [
                '    短期借款', '    交易性金融负债', '    应付票据', '    应付账款',
                '    预收款项', '    应付职工薪酬', '    应交税费', '    应付利息',
                '    应付股利', '    其他应付款', '    预提费用',
                '    一年内到期的非流动负债', '    其他流动负债'
            ])
            template['流动负债合计'].update({'本期': current_liabilities})
            
            # 计算非流动负债合计
            non_current_liabilities = sum(template[name]['本期'] for name in [
                '    长期借款', '    应付债券', '    长期应付款', '    专项应付款',
                '    预计负债', '    递延所得税负债', '    其他非流动负债'
            ])
            template['非流动负债合计'].update({'本期': non_current_liabilities})
            
            # 计算负债合计
            total_liabilities = current_liabilities + non_current_liabilities
            template['负债合计'].update({'本期': total_liabilities})
            
            # 计算所有者权益合计
            equity = sum(template[name]['本期'] for name in [
                '    股本', '    资本公积', '    盈余公积', '    未分配利润',
                '    少数股东权益', '    未结清对外担保余额'
            ])
            if '    减：库存股' in template:
                equity -= template['    减：库存股']['本期']
            template['所有者权益（或股东权益）合计'].update({'本期': equity})
            
            # 计算负债和所有者权益总计
            template['负债和所有者权益（或股东权益）总计'].update({'本期': total_liabilities + equity})
            
            # 对上期和年初数也进行同样的计算
            for period in ['上期', '年初']:
                self.calculate_period_totals(template, period)
                
        elif sheet_type == 'cash_flow':
            # 计算经营活动现金流入小计
            operating_inflow = sum(template[name][period] for name in [
                '    销售商品、提供劳务收到的现金', '    收到的税费返还',
                '    收到其他与经营活动有关的现金'
            ] for period in ['本期', '上期', '年初'])
            template['    经营活动现金流入小计'].update({'本期': operating_inflow})
            
            # 继续计算其他现金流量表的合计项...
            
        elif sheet_type == 'income_statement':
            # 计算营业利润
            operating_profit = (
                template['一、营业总收入']['本期'] -
                template['    减：营业成本']['本期'] -
                template['    营业税金及附加']['本期'] -
                template['    销售费用']['本期'] -
                template['    管理费用']['本期'] -
                template['    财务费用（收益以"－"号填列）']['本期'] -
                template['    资产减值损失']['本期'] +
                template['    加：公允价值变动净收益（净损失以"－"号填列）']['本期'] +
                template['    投资收益（净损失以"－"号填列）']['本期']
            )
            template['二、营业利润（亏损以"－"填列）'].update({'本期': operating_profit})
            
            # 继续计算其他损益表的合计项...
    
    def calculate_period_totals(self, template, period):
        """计算指定期间的合计项"""
        # 实现与本期计算相同的逻辑，但针对指定期间
        pass
    
    def calculate_financial_indicators(self, bs, is_, cf):
        """计算重点财务指标"""
        indicators = {
            '本期': {},
            '上期': {},
            '年初': {}
        }
        
        for period in ['本期', '上期', '年初']:
            try:
                # 1. 资产负债率（%）
                if bs['资产总计'].get(period, 0) != 0:
                    indicators[period]['资产负债率'] = \
                        (bs['负债合计'].get(period, 0) / bs['资产总计'].get(period, 0)) * 100
                
                # 2. 流动比率（%）
                if bs['流动负债合计'].get(period, 0) != 0:
                    indicators[period]['流动比率'] = \
                        (bs['流动资产合计'].get(period, 0) / bs['流动负债合计'].get(period, 0)) * 100
                
                # 3. 速动比率（%）
                quick_assets = bs['流动资产合计'].get(period, 0) \
                             - bs['    存货'].get(period, 0) \
                             - bs['    预付款项'].get(period, 0) \
                             - bs.get('    待摊费用', {}).get(period, 0)
                             
                if bs['流动负债合计'].get(period, 0) != 0:
                    indicators[period]['速动比率'] = \
                        (quick_assets / bs['流动负债合计'].get(period, 0)) * 100
                
                # 4. 利息保障倍数
                finance_cost = is_['    财务费用（收益以"－"号填列）'].get(period, 0)
                if finance_cost != 0:
                    total_profit = is_['三、利润总额（亏损总额以"－"填列）'].get(period, 0)
                    indicators[period]['利息保障倍数'] = (total_profit + finance_cost) / finance_cost
                
                # 5. 总资产周转率（次）
                if period != '年初':
                    avg_total_assets = (bs['资产总计'].get(period, 0) + bs['资产总计'].get('年初', 0)) / 2
                    if avg_total_assets != 0:
                        indicators[period]['总资产周转率'] = \
                            is_['一、营业总收入'].get(period, 0) / avg_total_assets
                
                # 6. 净资产收益率（%）
                if bs['所有者权益（或股东权益）合计'].get(period, 0) != 0:
                    indicators[period]['净资产收益率'] = \
                        (is_['四、净利润（净亏损以"－"号填列）'].get(period, 0) / 
                         bs['所有者权益（或股东权益）合计'].get(period, 0)) * 100
                
                # 7. 销售利润率（%）
                if is_['一、营业总收入'].get(period, 0) != 0:
                    indicators[period]['销售利润率'] = \
                        (is_['二、营业利润（亏损以"－"填列）'].get(period, 0) / 
                         is_['一、营业总收入'].get(period, 0)) * 100
                
                # 8. 经营活动现金流量/销售收入
                if is_['一、营业总收入'].get(period, 0) != 0:
                    indicators[period]['经营活动现金流量/销售收入'] = \
                        cf['    经营活动产生的现金流量净额'].get(period, 0) / \
                        is_['一、营业总收入'].get(period, 0)
                
                # 9. 存货周转率（次）
                if period != '年初':
                    avg_inventory = (bs['    存货'].get(period, 0) + bs['    存货'].get('年初', 0)) / 2
                    if avg_inventory != 0:
                        indicators[period]['存货周转率'] = \
                            is_['    减：营业成本'].get(period, 0) / avg_inventory
                
                # 10. 应收账款周转率（次）
                if period != '年初':
                    avg_accounts_receivable = (bs['    应收账款'].get(period, 0) + 
                                             bs['    应收账款'].get('年初', 0)) / 2
                    if avg_accounts_receivable != 0:
                        indicators[period]['应收账款周转率'] = \
                            is_['一、营业总收入'].get(period, 0) / avg_accounts_receivable
                
                # 11. 销售增长率（%）
                if period == '本期' and is_['一、营业总收入'].get('上期', 0) != 0:
                    indicators[period]['销售增长率'] = \
                        ((is_['一、营业总收入'].get(period, 0) - 
                          is_['一、营业总收入'].get('上期', 0)) / 
                         is_['一、营业总收入'].get('上期', 0)) * 100
                
                # 12. 总资产增长率（%）
                if period == '本期' and bs['资产总计'].get('上期', 0) != 0:
                    indicators[period]['总资产增长率'] = \
                        ((bs['资产总计'].get(period, 0) - bs['资产总计'].get('上期', 0)) / 
                         bs['资产总计'].get('上期', 0)) * 100
                
                # 13. 资本积累率（%）
                if period == '本期' and bs['所有者权益（或股东权益）合计'].get('上期', 0) != 0:
                    indicators[period]['资本积累率'] = \
                        ((bs['所有者权益（或股东权益）合计'].get(period, 0) - 
                          bs['所有者权益（或股东权益）合计'].get('上期', 0)) / 
                         bs['所有者权益（或股东权益）合计'].get('上期', 0)) * 100
                
                # 14. 总资产报酬率（%）
                if bs['资产总计'].get(period, 0) != 0:
                    indicators[period]['总资产报酬率'] = \
                        ((is_['三、利润总额（亏损总额以"－"填列）'].get(period, 0) + 
                          is_['    财务费用（收益以"－"号填列）'].get(period, 0)) / 
                         bs['资产总计'].get(period, 0)) * 100
                
                # 15. 成本费用利润率（%）
                total_cost = (is_['    减：营业成本'].get(period, 0) +
                            is_['    营业税金及附加'].get(period, 0) +
                            is_['    销售费用'].get(period, 0) +
                            is_['    管理费用'].get(period, 0) +
                            is_['    财务费用（收益以"－"号填列）'].get(period, 0))
                if total_cost != 0:
                    indicators[period]['成本费用利润率'] = \
                        (is_['二、营业利润（亏损以"－"填列）'].get(period, 0) / total_cost) * 100
                
                # 16. 总债务/EBITDA
                ebitda = (is_['三、利润总额（亏损总额以"－"填列）'].get(period, 0) + 
                         is_['    财务费用（收益以"－"号填列）'].get(period, 0) +
                         bs.get('    固定资产折旧', {}).get(period, 0) +
                         bs.get('    无形资产摊销', {}).get(period, 0) +
                         bs.get('    长期待摊费用摊销', {}).get(period, 0))
                if ebitda != 0:
                    indicators[period]['总债务/EBITDA'] = \
                        bs['负债合计'].get(period, 0) / ebitda
                
                # 17. 全部资本化比率（%）
                total_borrowings = (bs['    短期借款'].get(period, 0) + 
                                  bs['    长期借款'].get(period, 0))
                denominator = total_borrowings + bs['所有者权益（或股东权益）合计'].get(period, 0)
                if denominator != 0:
                    indicators[period]['全部资本化比率'] = \
                        (total_borrowings / denominator) * 100
                
                # 18. 经营活动现金净流量/总债务
                if bs['负债合计'].get(period, 0) != 0:
                    indicators[period]['经营活动现金净流量/总债务'] = \
                        cf['    经营活动产生的现金流量净额'].get(period, 0) / \
                        bs['负债合计'].get(period, 0)
                
                # 19. 流动资产周转率（次）
                if period != '年初':
                    avg_current_assets = (bs['流动资产合计'].get(period, 0) + 
                                        bs['流动资产合计'].get('年初', 0)) / 2
                    if avg_current_assets != 0:
                        indicators[period]['流动资产周转率'] = \
                            is_['一、营业总收入'].get(period, 0) / avg_current_assets
                
                # 20. 已获利息倍数
                if is_['    财务费用（收益以"－"号填列）'].get(period, 0) != 0:
                    indicators[period]['已获利息倍数'] = \
                        ((is_['三、利润总额（亏损总额以"－"填列）'].get(period, 0) + 
                          is_['    财务费用（收益以"－"号填列）'].get(period, 0)) / 
                         is_['    财务费用（收益以"－"号填列）'].get(period, 0))
                
                # 21. 营业收入现金含量（%）
                if is_['一、营业总收入'].get(period, 0) != 0:
                    indicators[period]['营业收入现金含量'] = \
                        (cf['    销售商品、提供劳务收到的现金'].get(period, 0) / 
                         is_['一、营业总收入'].get(period, 0)) * 100
                
                # 22. 经营活动现金净流量/流动负债（%）
                if bs['流动负债合计'].get(period, 0) != 0:
                    indicators[period]['经营活动现金净流量/流动负债'] = \
                        (cf['    经营活动产生的现金流量净额'].get(period, 0) / 
                         bs['流动负债合计'].get(period, 0)) * 100
                
            except Exception as e:
                self.log_message(f"计算{period}指标时出错：{str(e)}", "ERROR")
        
        return indicators
    
    def export_financial_indicators(self, workbook, indicators):
        """导出重点财务指标"""
        ws = workbook.create_sheet("重点财务指标")
        
        # 设置表头
        headers = ["指标", "本期", "上期", "年初"]
        for col, header in enumerate(headers, start=1):
            ws.cell(1, col, header)
        
        # 定义指标顺序和格式化
        indicator_formats = {
            "资产负债率": {"suffix": "%", "decimals": 2},
            "流动比率": {"suffix": "%", "decimals": 2},
            "总债务/EBITDA": {"suffix": "", "decimals": 2},
            "全部资本化比率": {"suffix": "%", "decimals": 2},
            "已获利息倍数": {"suffix": "", "decimals": 2},
            "速动比率": {"suffix": "%", "decimals": 2},
            "经营活动现金净流量/总债务": {"suffix": "", "decimals": 2},
            "净资产收益率": {"suffix": "%", "decimals": 2},
            "销售利润率": {"suffix": "%", "decimals": 2},
            "总资产收益率": {"suffix": "%", "decimals": 2},
            "经营活动现金流入量/销售收入": {"suffix": "", "decimals": 2},
            "成本费用利润率": {"suffix": "%", "decimals": 2},
            "总资产周转率": {"suffix": "次", "decimals": 2},
            "流动资产周转率": {"suffix": "次", "decimals": 2},
            "存货周转率": {"suffix": "次", "decimals": 2},
            "应收账款周转率": {"suffix": "次", "decimals": 2},
            "销售增长率": {"suffix": "%", "decimals": 2},
            "资本积累率": {"suffix": "%", "decimals": 2},
            "总资产增长率": {"suffix": "%", "decimals": 2},
            "销售利润率": {"suffix": "%", "decimals": 2},
            "总资产报酬率": {"suffix": "%", "decimals": 2},
            "利息保障倍数": {"suffix": "", "decimals": 2},
            "总债务/EBITDA": {"suffix": "", "decimals": 2},
            "全部资本化比率": {"suffix": "%", "decimals": 2},
            "经营活动现金净流量/总债务": {"suffix": "", "decimals": 2},
            "流动资产周转率": {"suffix": "次", "decimals": 2},
            "已获利息倍数": {"suffix": "", "decimals": 2},
            "营业收入现金含量": {"suffix": "%", "decimals": 2},
            "经营活动现金净流量/流动负债": {"suffix": "%", "decimals": 2}
        }
        
        # 写入数据
        row = 2
        for indicator, format_info in indicator_formats.items():
            # 写入指标名称
            ws.cell(row, 1, f"{indicator}({format_info['suffix']})" if format_info['suffix'] else indicator)
            
            # 写入各期间数据
            for col, period in enumerate(['本期', '上期', '年初'], start=2):
                value = indicators.get(period, {}).get(indicator)
                if isinstance(value, (int, float)):
                    formatted_value = round(value, format_info['decimals'])
                    ws.cell(row, col, formatted_value)
            row += 1
        
        # 调整列宽
        self.adjust_column_width(ws)
    
    def log_message(self, message, level="INFO"):
        """记录日志信息"""
        try:
            # 获取当前时间
            current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # 根据日志级别设置颜色
            colors = {
                "INFO": "black",
                "SUCCESS": "green",
                "WARNING": "orange",
                "ERROR": "red",
                "DEBUG": "blue"
            }
            
            # 格式化日志消息
            log_entry = f"[{current_time}] [{level}] {message}\n"
            
            # 在文本框中显示日志
            if hasattr(self, 'log_text') and self.log_text:
                # 临时启用编辑
                self.log_text.configure(state='normal')
                
                # 插入日志
                self.log_text.insert(tk.END, log_entry)
                
                # 应用颜色标签
                end_index = self.log_text.index(tk.END)
                start_index = f"{float(end_index) - 1} linestart"
                self.log_text.tag_add(level, start_index, f"{end_index} -1c")
                
                # 自动滚动到最新日志
                self.log_text.see(tk.END)
                
                # 恢复只读状态
                self.log_text.configure(state='disabled')
            
            # 同时输出到控制台（可选）
            # print(log_entry.strip())
            
        except Exception as e:
            # 如果日志记录失败，至少要在控制台显示错误
            print(f"日志记录失败：{str(e)}")
            print(f"原始消息：[{level}] {message}")

    def check_time_lock(self):
        """检查时间锁"""
        try:
            current_date = datetime.datetime.now().date()
            start_date = datetime.date(2025, 2, 5)
            end_date = datetime.date(2025, 6, 1)
            
            return start_date <= current_date <= end_date
        except:
            return False
    
    def run(self):
        """运行主程序"""
        self.root.mainloop()

    def names_match(self, name1, name2):
        """比较两个名称是否匹配"""
        # 移除所有空格和特殊字符后比较
        name1 = self.clean_item_name(name1)
        name2 = self.clean_item_name(name2)
        
        # 直接相等
        if name1 == name2:
            return True
            
        # 忽略"合计"、"小计"等后缀
        suffixes = ['合计', '小计', '总计', '净额', '净值', '：', ':', '总额']
        for suffix in suffixes:
            if name1.endswith(suffix):
                name1 = name1[:-len(suffix)]
            if name2.endswith(suffix):
                name2 = name2[:-len(suffix)]
                
        # 处理特殊的匹配规则
        special_matches = {
            '待摊费用': ['长期待摊费用'],
            '其它长期资产': ['其他非流动资产'],
            '所有者权益': ['所有者权益（或股东权益）'],
            '股东权益': ['所有者权益（或股东权益）'],
            '流动资产': ['流动资产：', '流动资产合计'],
            '非流动资产': ['非流动资产：', '非流动资产合计'],
            '流动负债': ['流动负债：', '流动负债合计'],
            '非流动负债': ['非流动负债：', '非流动负债合计'],
            '资产': ['资产总计', '资产合计'],
            '负债': ['负债合计', '负债总计'],
            '所有者权益': ['所有者权益（或股东权益）：', '所有者权益（或股东权益）合计']
        }
        
        # 检查特殊匹配
        for key, values in special_matches.items():
            if name1 == key and name2 in values:
                return True
            if name2 == key and name1 in values:
                return True
                
        # 如果都不匹配，返回标准比较结果
        return name1 == name2

    def get_synonyms(self):
        """获取同义词字典"""
        if not hasattr(self, '_synonyms'):
            self._synonyms = {
                # 流动资产类
                '货币资金': ['货币资金', '现金', '银行存款', '库存现金', '银行存款', '货币', '现金及存放中央银行款项'],
                '交易性金融资产': ['交易性金融资产', '交易性金融资产净额', 
                          '以公允价值计量且其变动计入当期损益的金融资产', '交易性投资'],
                '应收票据': ['应收票据', '应收票据净额', '应收票据及应收账款', '应收票据和应收账款'],
                '应收账款': ['应收账款', '应收账款净额', '应收款项', '应收票据及应收账款', '应收款'],
                '预付款项': ['预付款项', '预付账款', '预付款', '预付', '预付款项净额'],
                '应收利息': ['应收利息', '应收利息净额', '应收利息收入', '应收利息及应收股利'],
                '应收股利': ['应收股利', '应收股息', '应收股息红利', '应收利息及应收股利'],
                '其他应收款': ['其他应收款', '其他应收款净额', '其它应收款', '其它应收', '其他应收'],
                '存货': ['存货', '存货净额', '库存商品', '存货及合同履约成本', '库存'],
                '一年内到期的非流动资产': ['一年内到期的非流动资产', '一年内到期非流动资产', '一年内到期长期资产'],
                '其他流动资产': ['其他流动资产', '其它流动资产', '其他流动', '其它流动'],
                
                # 非流动资产类
                '可供出售金融资产': ['可供出售金融资产', '可供出售的金融资产', '可供出售投资'],
                '持有至到期投资': ['持有至到期投资', '持有到期投资', '持有至到期'],
                '长期应收款': ['长期应收款', '长期应收款项', '长期应收'],
                '长期股权投资': ['长期股权投资', '长期投资', '长期股权'],
                '投资性房地产': ['投资性房地产', '投资性房产', '投资房地产'],
                '固定资产': ['固定资产', '固定资产净额', '固定资产净值', '固定资产原价', '固定资产价值'],
                '在建工程': ['在建工程', '在建工程净额', '在建项目', '在建'],
                '工程物资': ['工程物资', '工程材料', '工程用料'],
                '固定资产清理': ['固定资产清理', '固定资产清算', '资产清理'],
                '生产性生物资产': ['生产性生物资产', '生物资产', '生产性生物'],
                '油气资产': ['油气资产', '石油天然气资产', '油气'],
                '无形资产': ['无形资产', '无形资产净额', '无形资产价值', '无形'],
                '开发支出': ['开发支出', '研发支出', '开发成本', '研发费用'],
                '商誉': ['商誉', '商誉净额', '商誉价值'],
                '长期待摊费用': ['长期待摊费用', '待摊费用', '长期待摊', '待摊'],
                '递延所得税资产': ['递延所得税资产', '递延税款', '递延所得税', '递延税资产'],
                '其他非流动资产': ['其他非流动资产', '其它非流动资产', '其他长期资产', '其它长期资产'],
                
                # 流动负债类
                '短期借款': ['短期借款', '短期贷款', '短期债务', '短期融资'],
                '交易性金融负债': ['交易性金融负债', '以公允价值计量且其变动计入当期损益的金融负债', '交易性负债'],
                '应付票据': ['应付票据', '应付汇票', '应付票据及应付账款'],
                '应付账款': ['应付账款', '应付账款净额', '应付款项', '应付票据及应付账款'],
                '预收款项': ['预收款项', '预收账款', '预收款', '合同负债', '预收'],
                '应付职工薪酬': ['应付职工薪酬', '应付工资', '工资福利', '应付工资薪酬'],
                '应交税费': ['应交税费', '应交税金', '应缴税金', '应交税款'],
                '应付利息': ['应付利息', '应付利息费用', '应付利息及应付股利'],
                '应付股利': ['应付股利', '应付股息', '应付利息及应付股利'],
                '其他应付款': ['其他应付款', '其他应付款净额', '其它应付款', '其他应付'],
                '一年内到期的非流动负债': ['一年内到期的非流动负债', '一年内到期非流动负债', '一年内到期长期负债'],
                '其他流动负债': ['其他流动负债', '其它流动负债', '其他流动', '其它流动'],
                
                # 非流动负债类
                '长期借款': ['长期借款', '长期贷款', '长期债务', '长期融资'],
                '应付债券': ['应付债券', '应付债券净额', '债券'],
                '长期应付款': ['长期应付款', '长期应付款项', '长期应付'],
                '专项应付款': ['专项应付款', '专项款项', '专项应付'],
                '预计负债': ['预计负债', '预计债务', '预提负债'],
                '递延所得税负债': ['递延所得税负债', '递延税负债', '递延税款负债'],
                '其他非流动负债': ['其他非流动负债', '其它非流动负债', '其他长期负债'],
                
                # 所有者权益类
                '股本': ['股本', '实收资本', '实收资本(或股本)', '注册资本', '股本金'],
                '资本公积': ['资本公积', '资本公积金', '资本溢价', '股本溢价'],
                '减：库存股': ['减：库存股', '库存股', '库存股份', '减库存股'],
                '盈余公积': ['盈余公积', '盈余公积金', '法定盈余', '盈余'],
                '未分配利润': ['未分配利润', '未分配利润(未弥补亏损)', '累计利润', '留存收益'],
                '少数股东权益': ['少数股东权益', '少数股东权益合计', '少数股东'],
                
                # 报表类别和合计项
                '流动资产': ['流动资产', '流动资产合计', '流动资产总计', '流动资产：', '流动资产总额'],
                '非流动资产': ['非流动资产', '非流动资产合计', '非流动资产总计', '非流动资产：', '非流动资产总额'],
                '资产总计': ['资产总计', '资产合计', '资产总额', '资产总额合计'],
                '流动负债': ['流动负债', '流动负债合计', '流动负债总计', '流动负债：', '流动负债总额'],
                '非流动负债': ['非流动负债', '非流动负债合计', '非流动负债总计', '非流动负债：', '非流动负债总额'],
                '负债合计': ['负债合计', '负债总计', '负债总额', '负债总额合计'],
                '所有者权益': ['所有者权益', '所有者权益（或股东权益）：', '所有者权益（或股东权益）合计', 
                        '股东权益', '股东权益合计', '所有者权益（或股东权益）总计', '所有者权益总额'],
                '负债和所有者权益总计': ['负债和所有者权益总计', '负债和所有者权益（或股东权益）总计', 
                             '负债及所有者权益总计', '负债和股东权益总计', '负债及股东权益总计'],
                             
                # 特殊项目
                '预提费用': ['预提费用', '预提成本费用', '预提支出'],
                '未结清对外担保余额': ['未结清对外担保余额', '对外担保余额', '担保余额']
            }
        return self._synonyms

if __name__ == "__main__":
    app = ReportConverter()
    app.run()
